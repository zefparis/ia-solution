"""
Module pour l'exportation rapide de documents vers différents formats
avec fonctionnalité d'aperçu
"""

import os
import json
import tempfile
from datetime import datetime
from flask import Blueprint, render_template, request, jsonify, send_file, url_for, flash, redirect, session
from pdfkit import from_string
import pdfkit
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import csv
import io
from auth import login_required
from models import db, User, Invoice, Quote, TaxReport
from models_business import BusinessReport
from flask import g
import language

# Fonction de traduction simplifiée
def _(key):
    """Traduction simplifiée pour ce module"""
    lang = getattr(g, 'lang', 'fr')
    return language.TRANSLATIONS.get(key, {}).get(lang, key)

# Configuration du Blueprint
export_bp = Blueprint('export', __name__)

# Formats d'exportation supportés
SUPPORTED_FORMATS = {
    'pdf': {
        'name': 'PDF',
        'icon': 'bi-file-earmark-pdf',
        'color': '#dc3545',
        'mime': 'application/pdf'
    },
    'excel': {
        'name': 'Excel',
        'icon': 'bi-file-earmark-excel',
        'color': '#198754',
        'mime': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    },
    'csv': {
        'name': 'CSV',
        'icon': 'bi-file-earmark-text',
        'color': '#0d6efd',
        'mime': 'text/csv'
    },
    'json': {
        'name': 'JSON',
        'icon': 'bi-file-earmark-code',
        'color': '#fd7e14',
        'mime': 'application/json'
    }
}

def init_app(app):
    """Initialiser les routes de l'exportation pour l'application Flask"""
    app.register_blueprint(export_bp, url_prefix='/export')

@export_bp.route('/')
def export_home():
    """Page d'accueil de l'exportation rapide"""
    # Vérifier si l'utilisateur est connecté
    username = session.get('username')
    user = None
    document_types = []
    
    if username:
        user = User.query.filter_by(username=username).first()
        
        # Récupérer les types de documents disponibles pour l'exportation
        # avec le nombre d'éléments disponibles par type
        if user:
            document_types = [
                {
                    'id': 'invoices',
                    'name': _('Factures'),
                    'count': Invoice.query.filter_by(user_id=user.id).count(),
                    'icon': 'bi-receipt',
                    'color': '#9c27b0'
                },
                {
                    'id': 'quotes',
                    'name': _('Devis'),
                    'count': Quote.query.filter_by(user_id=user.id).count(),
                    'icon': 'bi-clipboard',
                    'color': '#2196f3'
                },
                {
                    'id': 'tax_reports',
                    'name': _('Rapports fiscaux'),
                    'count': TaxReport.query.filter_by(user_id=user.id).count(),
                    'icon': 'bi-pie-chart',
                    'color': '#4caf50'
                },
                {
                    'id': 'business_reports',
                    'name': _('Analyses business'),
                    'count': BusinessReport.query.filter_by(user_id=user.id).count(),
                    'icon': 'bi-graph-up',
                    'color': '#ff9800'
                }
            ]
    
    # Passer la fonction de traduction au template
    return render_template(
        'export/index.html',
        user=user,
        document_types=document_types,
        formats=SUPPORTED_FORMATS,
        is_connected=(user is not None),
        _=_  # Passer la fonction de traduction
    )

@export_bp.route('/documents/<doc_type>')
@login_required
def list_documents(doc_type):
    """Liste les documents disponibles pour l'exportation selon le type"""
    username = session.get('username')
    user = User.query.filter_by(username=username).first()
    
    if doc_type == 'invoices':
        documents = Invoice.query.filter_by(user_id=user.id).order_by(Invoice.created_at.desc()).all()
        return render_template(
            'export/list_documents.html',
            doc_type=doc_type,
            doc_type_name=_('Factures'),
            documents=documents,
            formats=SUPPORTED_FORMATS,
            _=_  # Passer la fonction de traduction
        )
    
    elif doc_type == 'quotes':
        documents = Quote.query.filter_by(user_id=user.id).order_by(Quote.created_at.desc()).all()
        return render_template(
            'export/list_documents.html',
            doc_type=doc_type,
            doc_type_name=_('Devis'),
            documents=documents,
            formats=SUPPORTED_FORMATS,
            _=_  # Passer la fonction de traduction
        )
    
    elif doc_type == 'tax_reports':
        documents = TaxReport.query.filter_by(user_id=user.id).order_by(TaxReport.created_at.desc()).all()
        return render_template(
            'export/list_documents.html',
            doc_type=doc_type,
            doc_type_name=_('Rapports fiscaux'),
            documents=documents,
            formats=SUPPORTED_FORMATS,
            _=_  # Passer la fonction de traduction
        )
    
    elif doc_type == 'business_reports':
        documents = BusinessReport.query.filter_by(user_id=user.id).order_by(BusinessReport.created_at.desc()).all()
        return render_template(
            'export/list_documents.html',
            doc_type=doc_type,
            doc_type_name=_('Analyses business'),
            documents=documents,
            formats=SUPPORTED_FORMATS,
            _=_  # Passer la fonction de traduction
        )
    
    else:
        flash(_('Type de document non valide.'), 'danger')
        return redirect(url_for('export.export_home'))

@export_bp.route('/preview/<doc_type>/<int:doc_id>')
@login_required
def preview_document(doc_type, doc_id):
    """Affiche un aperçu du document à exporter"""
    username = session.get('username')
    user = User.query.filter_by(username=username).first()
    
    document = None
    template_name = None
    
    if doc_type == 'invoices':
        document = Invoice.query.filter_by(id=doc_id, user_id=user.id).first_or_404()
        template_name = 'export/preview_invoice.html'
    
    elif doc_type == 'quotes':
        document = Quote.query.filter_by(id=doc_id, user_id=user.id).first_or_404()
        template_name = 'export/preview_quote.html'
    
    elif doc_type == 'tax_reports':
        document = TaxReport.query.filter_by(id=doc_id, user_id=user.id).first_or_404()
        template_name = 'export/preview_tax_report.html'
    
    elif doc_type == 'business_reports':
        document = BusinessReport.query.filter_by(id=doc_id, user_id=user.id).first_or_404()
        template_name = 'export/preview_business_report.html'
    
    else:
        flash(_('Type de document non valide.'), 'danger')
        return redirect(url_for('export.export_home'))
    
    if not document:
        flash(_('Document non trouvé.'), 'danger')
        return redirect(url_for('export.list_documents', doc_type=doc_type))
    
    return render_template(
        template_name,
        document=document,
        formats=SUPPORTED_FORMATS,
        _=_  # Passer la fonction de traduction
    )

@export_bp.route('/generate/<doc_type>/<int:doc_id>/<format_type>')
@login_required
def generate_export(doc_type, doc_id, format_type):
    """Génère le document dans le format demandé et le retourne"""
    username = session.get('username')
    user = User.query.filter_by(username=username).first()
    
    # Vérifier que le format est supporté
    if format_type not in SUPPORTED_FORMATS:
        flash(_('Format non supporté.'), 'danger')
        return redirect(url_for('export.preview_document', doc_type=doc_type, doc_id=doc_id))
    
    # Récupérer le document selon son type
    document = None
    if doc_type == 'invoices':
        document = Invoice.query.filter_by(id=doc_id, user_id=user.id).first_or_404()
        filename = f"facture_{document.invoice_number}_{datetime.now().strftime('%Y%m%d')}"
    
    elif doc_type == 'quotes':
        document = Quote.query.filter_by(id=doc_id, user_id=user.id).first_or_404()
        filename = f"devis_{document.quote_number}_{datetime.now().strftime('%Y%m%d')}"
    
    elif doc_type == 'tax_reports':
        document = TaxReport.query.filter_by(id=doc_id, user_id=user.id).first_or_404()
        filename = f"rapport_fiscal_{document.id}_{datetime.now().strftime('%Y%m%d')}"
    
    elif doc_type == 'business_reports':
        document = BusinessReport.query.filter_by(id=doc_id, user_id=user.id).first_or_404()
        filename = f"analyse_business_{document.id}_{datetime.now().strftime('%Y%m%d')}"
    
    else:
        flash(_('Type de document non valide.'), 'danger')
        return redirect(url_for('export.export_home'))
    
    # Générer le contenu selon le format demandé
    if format_type == 'pdf':
        return generate_pdf(doc_type, document, filename)
    
    elif format_type == 'excel':
        return generate_excel(doc_type, document, filename)
    
    elif format_type == 'csv':
        return generate_csv(doc_type, document, filename)
    
    elif format_type == 'json':
        return generate_json(doc_type, document, filename)
    
    # Si on arrive ici, c'est qu'il y a un problème
    flash(_('Une erreur est survenue lors de la génération du document.'), 'danger')
    return redirect(url_for('export.preview_document', doc_type=doc_type, doc_id=doc_id))

def generate_pdf(doc_type, document, filename):
    """Génère un PDF à partir du document"""
    template_name = None
    
    if doc_type == 'invoices':
        template_name = 'export/pdf_templates/invoice.html'
    elif doc_type == 'quotes':
        template_name = 'export/pdf_templates/quote.html'
    elif doc_type == 'tax_reports':
        template_name = 'export/pdf_templates/tax_report.html'
    elif doc_type == 'business_reports':
        template_name = 'export/pdf_templates/business_report.html'
    
    # Rendre le template HTML
    html_content = render_template(template_name, document=document, _=_)
    
    # Créer un fichier temporaire pour le PDF
    pdf_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    pdf_file.close()
    
    # Options pour pdfkit
    options = {
        'page-size': 'A4',
        'margin-top': '0.75in',
        'margin-right': '0.75in',
        'margin-bottom': '0.75in',
        'margin-left': '0.75in',
        'encoding': 'UTF-8',
    }
    
    # Génération du PDF
    try:
        pdfkit.from_string(html_content, pdf_file.name, options=options)
        return send_file(
            pdf_file.name,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"{filename}.pdf"
        )
    except Exception as e:
        # En cas d'erreur, supprimer le fichier temporaire
        os.unlink(pdf_file.name)
        flash(f"Erreur lors de la génération du PDF: {str(e)}", 'danger')
        return redirect(url_for('export.preview_document', doc_type=doc_type, doc_id=document.id))

def generate_excel(doc_type, document, filename):
    """Génère un fichier Excel à partir du document"""
    wb = Workbook()
    ws = wb.active
    
    # Configuration selon le type de document
    if doc_type == 'invoices':
        # Configuration pour les factures
        ws.title = "Facture"
        headers = ["Description", "Quantité", "Prix unitaire", "TVA %", "Total HT", "Total TTC"]
        ws.append(headers)
        
        # Formater l'en-tête
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Ajouter les lignes de la facture
        row = 2
        for item in document.items:
            ws.append([
                item.description,
                item.quantity,
                item.unit_price,
                item.tax_rate,
                item.subtotal,
                item.total
            ])
            row += 1
        
        # Ajouter les totaux
        ws.append(["", "", "", "", "Total HT:", document.subtotal])
        ws.append(["", "", "", "", "Total TVA:", document.tax_amount])
        ws.append(["", "", "", "", "Total TTC:", document.total])
    
    elif doc_type == 'quotes':
        # Configuration similaire pour les devis
        ws.title = "Devis"
        headers = ["Description", "Quantité", "Prix unitaire", "TVA %", "Total HT", "Total TTC"]
        ws.append(headers)
        
        # Formater l'en-tête
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Ajouter les lignes du devis
        row = 2
        for item in document.items:
            ws.append([
                item.description,
                item.quantity,
                item.unit_price,
                item.tax_rate,
                item.subtotal,
                item.total
            ])
            row += 1
        
        # Ajouter les totaux
        ws.append(["", "", "", "", "Total HT:", document.subtotal])
        ws.append(["", "", "", "", "Total TVA:", document.tax_amount])
        ws.append(["", "", "", "", "Total TTC:", document.total])
    
    elif doc_type == 'tax_reports':
        # Configuration pour les rapports fiscaux
        ws.title = "Rapport Fiscal"
        headers = ["Catégorie", "Montant", "Type"]
        ws.append(headers)
        
        # Formater l'en-tête
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Ajouter les catégories de revenus
        ws.append(["Revenus", "", ""])
        for category, amount in document.income_categories.items():
            ws.append([category, amount, "Revenu"])
        
        # Ajouter les catégories de dépenses
        ws.append(["Dépenses", "", ""])
        for category, amount in document.expense_categories.items():
            ws.append([category, amount, "Dépense"])
        
        # Ajouter les totaux et statistiques
        ws.append(["", "", ""])
        ws.append(["Total Revenus", document.total_income, ""])
        ws.append(["Total Dépenses", document.total_expenses, ""])
        ws.append(["Bénéfice Net", document.net_income, ""])
        ws.append(["TVA Collectée", document.vat_collected, ""])
        ws.append(["TVA Déductible", document.vat_deductible, ""])
        ws.append(["TVA à Payer", document.vat_due, ""])
    
    elif doc_type == 'business_reports':
        # Configuration pour les rapports business
        ws.title = "Analyse Business"
        ws.append(["Entreprise", document.company_name])
        ws.append(["Secteur", document.sector])
        ws.append(["Date d'analyse", document.created_at.strftime('%d/%m/%Y')])
        ws.append(["", ""])
        
        # Ajouter l'analyse SWOT
        ws.append(["Analyse SWOT", ""])
        ws.append(["Forces", ""])
        for strength in document.swot_analysis.get('strengths', []):
            ws.append(["", strength])
        
        ws.append(["Faiblesses", ""])
        for weakness in document.swot_analysis.get('weaknesses', []):
            ws.append(["", weakness])
        
        ws.append(["Opportunités", ""])
        for opportunity in document.swot_analysis.get('opportunities', []):
            ws.append(["", opportunity])
        
        ws.append(["Menaces", ""])
        for threat in document.swot_analysis.get('threats', []):
            ws.append(["", threat])
        
        # Ajouter les recommandations
        ws.append(["", ""])
        ws.append(["Recommandations", ""])
        for recommendation in document.recommendations:
            ws.append(["", recommendation])
    
    # Ajuster la largeur des colonnes
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 5
    
    # Créer un buffer pour le fichier Excel
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Retourner le fichier Excel
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"{filename}.xlsx"
    )

def generate_csv(doc_type, document, filename):
    """Génère un fichier CSV à partir du document"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Configuration selon le type de document
    if doc_type == 'invoices':
        # En-tête pour les factures
        writer.writerow(["Facture", document.invoice_number])
        writer.writerow(["Date", document.issue_date.strftime('%d/%m/%Y')])
        writer.writerow(["Client", document.client.name if document.client else "N/A"])
        writer.writerow([])
        writer.writerow(["Description", "Quantité", "Prix unitaire", "TVA %", "Total HT", "Total TTC"])
        
        # Lignes de la facture
        for item in document.items:
            writer.writerow([
                item.description,
                item.quantity,
                item.unit_price,
                item.tax_rate,
                item.subtotal,
                item.total
            ])
        
        # Totaux
        writer.writerow([])
        writer.writerow(["", "", "", "", "Total HT:", document.subtotal])
        writer.writerow(["", "", "", "", "Total TVA:", document.tax_amount])
        writer.writerow(["", "", "", "", "Total TTC:", document.total])
    
    elif doc_type == 'quotes':
        # En-tête pour les devis
        writer.writerow(["Devis", document.quote_number])
        writer.writerow(["Date", document.issue_date.strftime('%d/%m/%Y')])
        writer.writerow(["Client", document.client.name if document.client else "N/A"])
        writer.writerow([])
        writer.writerow(["Description", "Quantité", "Prix unitaire", "TVA %", "Total HT", "Total TTC"])
        
        # Lignes du devis
        for item in document.items:
            writer.writerow([
                item.description,
                item.quantity,
                item.unit_price,
                item.tax_rate,
                item.subtotal,
                item.total
            ])
        
        # Totaux
        writer.writerow([])
        writer.writerow(["", "", "", "", "Total HT:", document.subtotal])
        writer.writerow(["", "", "", "", "Total TVA:", document.tax_amount])
        writer.writerow(["", "", "", "", "Total TTC:", document.total])
    
    elif doc_type == 'tax_reports':
        # En-tête pour les rapports fiscaux
        writer.writerow(["Rapport Fiscal", document.id])
        writer.writerow(["Période", f"{document.start_date.strftime('%d/%m/%Y')} - {document.end_date.strftime('%d/%m/%Y')}"])
        writer.writerow([])
        
        # Catégories de revenus
        writer.writerow(["Revenus"])
        writer.writerow(["Catégorie", "Montant"])
        for category, amount in document.income_categories.items():
            writer.writerow([category, amount])
        
        # Catégories de dépenses
        writer.writerow([])
        writer.writerow(["Dépenses"])
        writer.writerow(["Catégorie", "Montant"])
        for category, amount in document.expense_categories.items():
            writer.writerow([category, amount])
        
        # Totaux et statistiques
        writer.writerow([])
        writer.writerow(["Statistiques"])
        writer.writerow(["Total Revenus", document.total_income])
        writer.writerow(["Total Dépenses", document.total_expenses])
        writer.writerow(["Bénéfice Net", document.net_income])
        writer.writerow(["TVA Collectée", document.vat_collected])
        writer.writerow(["TVA Déductible", document.vat_deductible])
        writer.writerow(["TVA à Payer", document.vat_due])
    
    elif doc_type == 'business_reports':
        # En-tête pour les rapports business
        writer.writerow(["Analyse Business"])
        writer.writerow(["Entreprise", document.company_name])
        writer.writerow(["Secteur", document.sector])
        writer.writerow(["Date d'analyse", document.created_at.strftime('%d/%m/%Y')])
        writer.writerow([])
        
        # Analyse SWOT
        writer.writerow(["Analyse SWOT"])
        writer.writerow(["Forces"])
        for strength in document.swot_analysis.get('strengths', []):
            writer.writerow(["", strength])
        
        writer.writerow([])
        writer.writerow(["Faiblesses"])
        for weakness in document.swot_analysis.get('weaknesses', []):
            writer.writerow(["", weakness])
        
        writer.writerow([])
        writer.writerow(["Opportunités"])
        for opportunity in document.swot_analysis.get('opportunities', []):
            writer.writerow(["", opportunity])
        
        writer.writerow([])
        writer.writerow(["Menaces"])
        for threat in document.swot_analysis.get('threats', []):
            writer.writerow(["", threat])
        
        # Recommandations
        writer.writerow([])
        writer.writerow(["Recommandations"])
        for recommendation in document.recommendations:
            writer.writerow(["", recommendation])
    
    # Récupérer le contenu du buffer
    output.seek(0)
    
    # Retourner le fichier CSV
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f"{filename}.csv"
    )

def generate_json(doc_type, document, filename):
    """Génère un fichier JSON à partir du document"""
    data = {}
    
    # Configuration selon le type de document
    if doc_type == 'invoices':
        # Données pour les factures
        data = {
            'invoice_number': document.invoice_number,
            'issue_date': document.issue_date.strftime('%Y-%m-%d'),
            'due_date': document.due_date.strftime('%Y-%m-%d') if document.due_date else None,
            'client': {
                'id': document.client.id if document.client else None,
                'name': document.client.name if document.client else None,
                'email': document.client.email if document.client else None,
                'address': document.client.address if document.client else None
            },
            'items': [{
                'description': item.description,
                'quantity': float(item.quantity),
                'unit_price': float(item.unit_price),
                'tax_rate': float(item.tax_rate),
                'subtotal': float(item.subtotal),
                'total': float(item.total)
            } for item in document.items],
            'subtotal': float(document.subtotal),
            'tax_amount': float(document.tax_amount),
            'total': float(document.total),
            'notes': document.notes,
            'payment_info': document.payment_info,
            'status': document.status
        }
    
    elif doc_type == 'quotes':
        # Données pour les devis
        data = {
            'quote_number': document.quote_number,
            'issue_date': document.issue_date.strftime('%Y-%m-%d'),
            'valid_until': document.valid_until.strftime('%Y-%m-%d') if document.valid_until else None,
            'client': {
                'id': document.client.id if document.client else None,
                'name': document.client.name if document.client else None,
                'email': document.client.email if document.client else None,
                'address': document.client.address if document.client else None
            },
            'items': [{
                'description': item.description,
                'quantity': float(item.quantity),
                'unit_price': float(item.unit_price),
                'tax_rate': float(item.tax_rate),
                'subtotal': float(item.subtotal),
                'total': float(item.total)
            } for item in document.items],
            'subtotal': float(document.subtotal),
            'tax_amount': float(document.tax_amount),
            'total': float(document.total),
            'notes': document.notes,
            'conditions': document.conditions,
            'status': document.status
        }
    
    elif doc_type == 'tax_reports':
        # Données pour les rapports fiscaux
        data = {
            'id': document.id,
            'period': {
                'start_date': document.start_date.strftime('%Y-%m-%d'),
                'end_date': document.end_date.strftime('%Y-%m-%d')
            },
            'income_categories': document.income_categories,
            'expense_categories': document.expense_categories,
            'statistics': {
                'total_income': float(document.total_income),
                'total_expenses': float(document.total_expenses),
                'net_income': float(document.net_income),
                'profit_margin': document.profit_margin,
                'vat_collected': float(document.vat_collected),
                'vat_deductible': float(document.vat_deductible),
                'vat_due': float(document.vat_due)
            },
            'notes': document.notes,
            'ai_analysis': document.ai_analysis
        }
    
    elif doc_type == 'business_reports':
        # Données pour les rapports business
        data = {
            'id': document.id,
            'company_name': document.company_name,
            'sector': document.sector,
            'creation_date': document.created_at.strftime('%Y-%m-%d'),
            'swot_analysis': document.swot_analysis,
            'recommendations': document.recommendations,
            'business_plan': document.business_plan,
            'market_analysis': document.market_analysis
        }
    
    # Convertir les données en JSON
    json_data = json.dumps(data, ensure_ascii=False, indent=4)
    
    # Retourner le fichier JSON
    return send_file(
        io.BytesIO(json_data.encode('utf-8')),
        mimetype='application/json',
        as_attachment=True,
        download_name=f"{filename}.json"
    )